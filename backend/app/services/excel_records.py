from __future__ import annotations

import json
import os
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import DEFAULT_EXCEL_PATH
from ..models import ActivityEvent, Employee, LeaveBalance, LeaveRequest, PolicyDate, RequestDecision


SHEETS = ["Employees", "LeaveRequests", "Decisions", "Balances", "PolicyDates", "ActivityEvents"]
HEADERS = {
    "Employees": ["timestamp", "action", "actor", "employee_id", "name", "email", "role", "department", "payload"],
    "LeaveRequests": ["timestamp", "action", "actor", "request_id", "employee_id", "status", "payload"],
    "Decisions": ["timestamp", "action", "actor", "request_id", "note", "payload"],
    "Balances": ["timestamp", "action", "actor", "employee_id", "leave_type", "used", "payload"],
    "PolicyDates": ["timestamp", "action", "actor", "policy_date", "kind", "label", "payload"],
    "ActivityEvents": ["timestamp", "action", "actor", "kind", "text", "payload"],
}


last_error: str | None = None


def workbook_path() -> Path:
    path = Path(os.environ.get("DAYBOOK_EXCEL_PATH", str(DEFAULT_EXCEL_PATH)))
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def ensure_workbook(path: Path | None = None) -> Path:
    path = path or workbook_path()
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        default = wb.active
        wb.remove(default)
    for name in SHEETS:
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            ws.append(HEADERS[name])
    wb.save(path)
    wb.close()
    return path


def snapshot(payload: dict[str, Any]) -> str:
    return json.dumps(payload, default=str, sort_keys=True)


def append_record(sheet: str, action: str, payload: dict[str, Any], actor: str | None = None) -> None:
    global last_error
    try:
        path = ensure_workbook()
        wb = load_workbook(path)
        ws = wb[sheet]
        now = datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds")
        if sheet == "Employees":
            ws.append([now, action, actor, payload.get("id"), payload.get("name"), payload.get("email"), payload.get("role"), payload.get("department"), snapshot(payload)])
        elif sheet == "LeaveRequests":
            ws.append([now, action, actor, payload.get("id"), payload.get("employee_id"), payload.get("status"), snapshot(payload)])
        elif sheet == "Decisions":
            ws.append([now, action, actor, payload.get("request_id"), payload.get("note"), snapshot(payload)])
        elif sheet == "Balances":
            ws.append([now, action, actor, payload.get("employee_id"), payload.get("leave_type"), payload.get("used"), snapshot(payload)])
        elif sheet == "PolicyDates":
            ws.append([now, action, actor, payload.get("date"), payload.get("kind"), payload.get("label"), snapshot(payload)])
        elif sheet == "ActivityEvents":
            ws.append([now, action, actor, payload.get("kind"), payload.get("text"), snapshot(payload)])
        wb.save(path)
        wb.close()
        last_error = None
    except Exception as exc:  # pragma: no cover - intentionally defensive for admin status.
        last_error = str(exc)


def status() -> dict[str, Any]:
    path = ensure_workbook()
    wb = load_workbook(path)
    payload = {
        "path": str(path),
        "exists": path.exists(),
        "sheets": {name: max(0, wb[name].max_row - 1) for name in SHEETS},
        "last_error": last_error,
    }
    wb.close()
    return payload


def rebuild_from_db(db: Session) -> dict[str, Any]:
    path = workbook_path()
    if path.exists():
        path.unlink()
    ensure_workbook(path)
    for employee in db.query(Employee).all():
        append_record("Employees", "rebuild", employee_payload(employee), None)
    for balance in db.query(LeaveBalance).all():
        append_record("Balances", "rebuild", balance_payload(balance), None)
    for request in db.query(LeaveRequest).all():
        append_record("LeaveRequests", "rebuild", request_payload(request), None)
    for decision in db.query(RequestDecision).all():
        append_record("Decisions", "rebuild", decision_payload(decision), decision.actor_id)
    for policy_date in db.query(PolicyDate).all():
        append_record("PolicyDates", "rebuild", policy_payload(policy_date), None)
    for event in db.query(ActivityEvent).all():
        append_record("ActivityEvents", "rebuild", activity_payload(event), event.actor_id)
    return status()


def employee_payload(item: Employee) -> dict[str, Any]:
    return {
        "id": item.id,
        "name": item.name,
        "email": item.email,
        "role": item.role,
        "title": item.title,
        "department": item.department,
        "manager_id": item.manager_id,
        "join_date": item.join_date,
        "phone": item.phone,
        "is_active": item.is_active,
    }


def balance_payload(item: LeaveBalance) -> dict[str, Any]:
    return {"employee_id": item.employee_id, "leave_type": item.leave_type, "used": item.used}


def request_payload(item: LeaveRequest) -> dict[str, Any]:
    return {
        "id": item.id,
        "employee_id": item.employee_id,
        "leave_type": item.leave_type,
        "start_date": item.start_date,
        "end_date": item.end_date,
        "half_day": item.half_day,
        "days": item.days,
        "status": item.status,
        "reason": item.reason,
        "contact": item.contact,
        "handover": item.handover,
        "applied_on": item.applied_on,
        "decided_by": item.decided_by,
        "decision_reason": item.decision_reason,
    }


def decision_payload(item: RequestDecision) -> dict[str, Any]:
    return {"request_id": item.request_id, "actor_id": item.actor_id, "action": item.action, "note": item.note, "created_at": item.created_at}


def policy_payload(item: PolicyDate) -> dict[str, Any]:
    return {"date": item.date, "kind": item.kind, "label": item.label, "blocks_leave": item.blocks_leave}


def activity_payload(item: ActivityEvent) -> dict[str, Any]:
    return {"kind": item.kind, "text": item.text, "actor_id": item.actor_id, "created_at": item.created_at}
