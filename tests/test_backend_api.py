from __future__ import annotations

import os
import sys
import tempfile
import unittest
from json import loads
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


class BackendApiTests(unittest.TestCase):
    def setUp(self):
        try:
            import fastapi  # noqa: F401
            import openpyxl  # noqa: F401
            import sqlalchemy  # noqa: F401
        except ImportError:
            self.skipTest("FastAPI, SQLAlchemy, and openpyxl are not installed. Run pip install -r requirements.txt.")

        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "daybook.sqlite3"
        self.excel_path = Path(self.temp_dir.name) / "daybook_records.xlsx"
        self.original_db_url = os.environ.get("DAYBOOK_DATABASE_URL")
        self.original_excel_path = os.environ.get("DAYBOOK_EXCEL_PATH")
        os.environ["DAYBOOK_DATABASE_URL"] = f"sqlite:///{self.db_path.as_posix()}"
        os.environ["DAYBOOK_EXCEL_PATH"] = str(self.excel_path)

        modules = [name for name in list(sys.modules) if name.startswith("backend.app")]
        for name in modules:
            sys.modules.pop(name, None)

        from fastapi.testclient import TestClient
        from backend.app.seed import seed
        from backend.app.main import app

        seed(reset=True)
        self.client = TestClient(app)

    def tearDown(self):
        for module_name in ("backend.app.database",):
            module = sys.modules.get(module_name)
            if module is not None:
                module.engine.dispose()
        if self.original_db_url is None:
            os.environ.pop("DAYBOOK_DATABASE_URL", None)
        else:
            os.environ["DAYBOOK_DATABASE_URL"] = self.original_db_url
        if self.original_excel_path is None:
            os.environ.pop("DAYBOOK_EXCEL_PATH", None)
        else:
            os.environ["DAYBOOK_EXCEL_PATH"] = self.original_excel_path
        self.temp_dir.cleanup()

    def headers_for(self, email: str, password: str) -> dict[str, str]:
        response = self.client.post("/auth/login", json={"email": email, "password": password})
        self.assertEqual(response.status_code, 200, response.text)
        return {"X-User-Id": response.json()["token"]}

    def test_login_dashboard_and_health(self):
        headers = self.headers_for("maya.lin@acme.co", "employee-demo")

        self.assertEqual(self.client.get("/health").json()["status"], "ok")
        dashboard = self.client.get("/analytics/dashboard", headers=headers)

        self.assertEqual(dashboard.status_code, 200, dashboard.text)
        self.assertEqual(dashboard.json()["user"]["email"], "maya.lin@acme.co")

    def test_hr_can_add_employee_and_excel_status_updates(self):
        headers = self.headers_for("grace.hall@acme.co", "hr-demo")

        response = self.client.post(
            "/employees",
            headers=headers,
            json={
                "name": "Nina Patel",
                "email": "nina.patel@acme.co",
                "role": "employee",
                "title": "Designer",
                "department": "Engineering",
                "manager_id": "e5",
                "join_date": "2026-01-10",
                "phone": "+1 415 555 0999",
                "is_active": True,
                "balances": {"annual": 1, "earned": 0, "sick": 0},
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["email"], "nina.patel@acme.co")
        status = self.client.get("/records/excel/status", headers=headers).json()
        self.assertGreaterEqual(status["sheets"]["Employees"], 1)
        self.assertGreaterEqual(status["sheets"]["Balances"], 1)

    def test_hr_can_add_manager_then_employee_and_excel_reflects_both(self):
        from openpyxl import load_workbook

        headers = self.headers_for("grace.hall@acme.co", "hr-demo")

        manager = self.client.post(
            "/employees",
            headers=headers,
            json={
                "name": "Zara Khan",
                "email": "zara.khan@acme.co",
                "role": "manager",
                "title": "Support Manager",
                "department": "People",
                "manager_id": "e7",
                "join_date": "2026-02-01",
                "phone": "+1 415 555 0210",
                "is_active": True,
                "balances": {"annual": 0, "earned": 0, "sick": 0},
            },
        )
        self.assertEqual(manager.status_code, 200, manager.text)
        self.assertEqual(manager.json()["role"], "manager")
        manager_id = manager.json()["id"]

        employee = self.client.post(
            "/employees",
            headers=headers,
            json={
                "name": "Omar Silva",
                "email": "omar.silva@acme.co",
                "role": "employee",
                "title": "People Operations Associate",
                "department": "People",
                "manager_id": manager_id,
                "join_date": "2026-03-15",
                "phone": "+1 415 555 0211",
                "is_active": True,
                "balances": {"annual": 2, "earned": 1, "sick": 0},
            },
        )
        self.assertEqual(employee.status_code, 200, employee.text)
        self.assertEqual(employee.json()["manager_name"], "Zara Khan")

        dashboard = self.client.get("/analytics/dashboard", headers=headers)
        self.assertEqual(dashboard.status_code, 200, dashboard.text)
        employees = {item["email"]: item for item in dashboard.json()["employees"]}
        self.assertEqual(employees["zara.khan@acme.co"]["role"], "manager")
        self.assertEqual(employees["omar.silva@acme.co"]["manager_id"], manager_id)
        self.assertEqual(employees["omar.silva@acme.co"]["manager_name"], "Zara Khan")

        status = self.client.get("/records/excel/status", headers=headers).json()
        self.assertGreaterEqual(status["sheets"]["Employees"], 9)
        self.assertGreaterEqual(status["sheets"]["Balances"], 54)

        workbook = load_workbook(self.excel_path)
        try:
            rows = list(workbook["Employees"].iter_rows(values_only=True))
        finally:
            workbook.close()
        employee_rows = {row[4]: row for row in rows[1:]}
        self.assertIn("Maya Lin", employee_rows)
        self.assertEqual(employee_rows["Zara Khan"][6], "manager")
        self.assertEqual(employee_rows["Omar Silva"][6], "employee")
        self.assertEqual(loads(employee_rows["Omar Silva"][8])["manager_id"], manager_id)

    def test_hr_cannot_create_another_hr_user(self):
        headers = self.headers_for("grace.hall@acme.co", "hr-demo")

        response = self.client.post(
            "/employees",
            headers=headers,
            json={
                "name": "Admin Copy",
                "email": "admin.copy@acme.co",
                "role": "hr",
                "title": "HR Admin",
                "department": "People",
                "manager_id": None,
                "join_date": "2026-01-10",
                "phone": "+1 415 555 0998",
                "is_active": True,
                "balances": {},
            },
        )

        self.assertEqual(response.status_code, 422)

    def test_employee_request_and_manager_approval(self):
        employee_headers = self.headers_for("maya.lin@acme.co", "employee-demo")
        manager_headers = self.headers_for("david.okafor@acme.co", "manager-demo")

        response = self.client.post(
            "/leave-requests",
            headers=employee_headers,
            json={
                "leave_type": "casual",
                "start_date": "2026-07-06",
                "end_date": "2026-07-06",
                "half_day": False,
                "reason": "Appointment",
            },
        )

        self.assertEqual(response.status_code, 200, response.text)
        request_id = response.json()["id"]
        approval = self.client.post(f"/leave-requests/{request_id}/approve", headers=manager_headers, json={"note": "Enjoy"})

        self.assertEqual(approval.status_code, 200, approval.text)
        self.assertEqual(approval.json()["status"], "approved")


if __name__ == "__main__":
    unittest.main()
